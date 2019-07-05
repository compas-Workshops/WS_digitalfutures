from __future__ import print_function
from __future__ import absolute_import
from __future__ import division

import os

from openpyxl import Workbook

from compas.geometry import distance_point_point
from compas.geometry import add_vectors
from compas.geometry import intersection_line_plane
from compas.geometry import cross_vectors
from compas.geometry import subtract_vectors
from compas.geometry import normalize_vector
from compas.geometry import scale_vector

from compas_fofin.datastructures import Shell

# ==============================================================================
# Helpers
# ==============================================================================

def append_to_sheet(sheet, beamA, beamB):
    for u in beamA:
        nbrs = SHELL.vertex_neighbors(u)
        v = None
        for nbr in nbrs:
            if nbr in boundary:
                continue
            v = nbr
            break
        if v is None:
            continue
        edges = SHELL.get_continuous_edges((u, v), directed=False)
        last = edges[-1][1]
        if last in beamB:
            values = LENGTHS[u][:]
            values += [SHELL.edge_length(u, v) for u, v in edges]
            values += LENGTHS[last][::-1]
            accumulated = []
            value = 0
            for v in values:
                value += v
                accumulated.append(value)
            sheet.append([u] + [round(1e3 * v, 1) for v in accumulated])
        else:
            values = LENGTHS[u][:]
            values += [SHELL.edge_length(u, v) for u, v in edges[:-1]]
            values += [SHELL.edge_length(* edges[-1]) - 0.100]
            values += [0.100, 0.100]
            accumulated = []
            value = 0
            for v in values:
                value += v
                accumulated.append(value)
            sheet.append([u] + [round(1e3 * v, 1) for v in accumulated])


def beam_plane(beam):
    p1 = SHELL.vertex_coordinates(beam[1])
    p0 = SHELL.vertex_coordinates(beam[0])
    xaxis = subtract_vectors(p1, p0)
    yaxis = cross_vectors(ZAXIS, xaxis)
    normal = normalize_vector(yaxis)
    origin = add_vectors(p0, scale_vector(normal, OFFSET))
    return origin, normal

# ==============================================================================
# Initialise
# ==============================================================================

HERE = os.path.dirname(__file__)
DATA = os.path.abspath(os.path.join(HERE, '..', 'data'))
FILE_I = os.path.join(DATA, 'data.json')
FILE_O = os.path.join(DATA, 'data-fabrication-cables.xlsx')

SHELL = Shell.from_json(FILE_I)

# ==============================================================================
# Beam vertices
# ==============================================================================

BEAMS = [
    [268, 258, 115, 106, 114, 269, 281, 145],
    [57, 4, 259, 278, 54],
    [116, 261, 282, 60, 79, 52, 260, 5],
    [146, 270, 117, 107],
    [8, 144, 272, 175]]

OFFSET = 0.5
ZAXIS = [0, 0, 1.0]

NORTH = beam_plane(BEAMS[0])
WEST = beam_plane(BEAMS[1])
SOUTH = beam_plane(BEAMS[2])

PLANES = [
    NORTH,
    WEST,
    SOUTH,
    SOUTH,
    NORTH]

# ==============================================================================
# Connector cables
# ==============================================================================

LENGTHS = {}

for i, keys in enumerate(BEAMS):
    plane = PLANES[i]

    points = []
    for key in keys:
        a = SHELL.vertex_coordinates(key)
        r = SHELL.get_vertex_attributes(key, ['rx', 'ry', 'rz'])
        b = add_vectors(a, r)

        line = a, b
        x = intersection_line_plane(line, plane)
        direction = normalize_vector(subtract_vectors(a, x))

        points.append(x)

        x0 = x
        x1 = add_vectors(x0, scale_vector(direction, 0.040))
        x2 = add_vectors(x1, scale_vector(direction, 0.300))
        x3 = add_vectors(x2, scale_vector(direction, 0.100))
        x4 = a

        LENGTHS[key] = [
            0.1,
            distance_point_point(x2, x3),
            distance_point_point(x3, x4)]

# ==============================================================================
# Export
# ==============================================================================

wb = Workbook()

boundary = set(SHELL.vertices_on_boundary())

BEAMS = [116, 261, 282, 60, 79, 52, 260, 5]
beamW = [57, 4, 259, 278, 54]
beamN = [268, 258, 115, 106, 114, 269, 281, 145]
BEAMSE = [146, 270, 117, 107]
beamNE = [8, 144, 272, 175]

ws_S = wb.active
ws_S.title = "SOUTH"

ws_W = wb.create_sheet()
ws_W.title = "WEST"

ws_N = wb.create_sheet()
ws_N.title = "NORTH"

ws_E = wb.create_sheet()
ws_E.title = "EAST"

append_to_sheet(ws_S, BEAMS, beamN)
append_to_sheet(ws_W, beamW, [])
append_to_sheet(ws_N, beamN, BEAMS)
append_to_sheet(ws_E, BEAMSE, beamNE)

wb.save(FILE_O)
