from __future__ import print_function
from __future__ import absolute_import
from __future__ import division

import os

from openpyxl import Workbook

from compas.utilities import pairwise
from compas.geometry import distance_point_point
from compas.geometry import add_vectors
from compas.geometry import intersection_line_plane
from compas.geometry import Frame
from compas.geometry import bounding_box
from compas.geometry import bounding_box_xy
from compas.geometry import cross_vectors
from compas.geometry import subtract_vectors
from compas.geometry import normalize_vector
from compas.geometry import scale_vector
from compas.geometry import offset_polygon
from compas.rpc import Proxy

from compas_fofin.datastructures import Shell
from compas_fofin.rhino import ShellArtist

numerical = Proxy('compas.numerical')


def framelines(origin, xaxis, yaxis, zaxis, name):
    lines = []
    lines.append({
        'start' : origin,
        'end'   : add_vectors(origin, scale_vector(xaxis, 0.1)),
        'color' : (255, 0, 0),
        'name'  : "{}.X".format(name)
    })
    lines.append({
        'start' : origin,
        'end'   : add_vectors(origin, scale_vector(yaxis, 0.1)),
        'color' : (0, 255, 0),
        'name'  : "{}.Y".format(name)
    })
    lines.append({
        'start' : origin,
        'end'   : add_vectors(origin, scale_vector(zaxis, 0.1)),
        'color' : (0, 0, 255),
        'name'  : "{}.Z".format(name)
    })
    return lines


def append_to_sheet(sheet, beamA, beamB):
    for u in beamA:
        nbrs = shell.vertex_neighbors(u)
        v = None
        for nbr in nbrs:
            if nbr in boundary:
                continue
            v = nbr
            break
        if v is None:
            continue
        edges = shell.get_continuous_edges((u, v), directed=False)
        last = edges[-1][1]
        if last in beamB:
            values = LENGTHS[u][:]
            values += [shell.edge_length(u, v) for u, v in edges]
            values += LENGTHS[last][::-1]
            accumulated = []
            value = 0
            for v in values:
                value += v
                accumulated.append(value)
            sheet.append([u] + [round(1e3 * v, 1) for v in accumulated])
        else:
            values = LENGTHS[u][:]
            values += [shell.edge_length(u, v) for u, v in edges[:-1]]
            values += [shell.edge_length(* edges[-1]) - 0.100]
            values += [0.100, 0.100]
            accumulated = []
            value = 0
            for v in values:
                value += v
                accumulated.append(value)
            sheet.append([u] + [round(1e3 * v, 1) for v in accumulated])


# ==============================================================================
# Initialise
# ==============================================================================

HERE = os.path.dirname(__file__)
DATA = os.path.abspath(os.path.join(HERE, '..', 'data'))
FILE_I = os.path.join(DATA, 'data.json')
FILE_O = os.path.join(DATA, 'data-fabrication-cables.xlsx')

shell = Shell.from_json(FILE_I)

# ==============================================================================
# Beam vertices
# ==============================================================================

beams = [
    [268, 258, 115, 106, 114, 269, 281, 145],  # NORTH
    [57, 4, 259, 278, 54],  # WEST
    [116, 261, 282, 60, 79, 52, 260, 5],  # SOUTH
    [146, 270, 117, 107],  # SOUTH
    [8, 144, 272, 175],  # NORTH
]

offset = 0.5
zaxis = [0, 0, 1.0]

p1 = shell.vertex_coordinates(beams[0][1])
p0 = shell.vertex_coordinates(beams[0][0])
xaxis = subtract_vectors(p1, p0)
yaxis = cross_vectors(zaxis, xaxis)
normal = normalize_vector(yaxis)
origin = add_vectors(p0, scale_vector(normal, offset))
NORTH = (origin, normal)

p1 = shell.vertex_coordinates(beams[1][1])
p0 = shell.vertex_coordinates(beams[1][0])
xaxis = subtract_vectors(p1, p0)
yaxis = cross_vectors(zaxis, xaxis)
normal = normalize_vector(yaxis)
origin = add_vectors(p0, scale_vector(normal, offset))
WEST = (origin, normal)

p1 = shell.vertex_coordinates(beams[2][1])
p0 = shell.vertex_coordinates(beams[2][0])
xaxis = subtract_vectors(p1, p0)
yaxis = cross_vectors(zaxis, xaxis)
normal = normalize_vector(yaxis)
origin = add_vectors(p0, scale_vector(normal, offset))
SOUTH = (origin, normal)

planes = [
    NORTH,
    WEST,
    SOUTH,
    SOUTH,
    NORTH
]

# ==============================================================================
# Horizontal beams
# ==============================================================================

LINES = []
POINTS = []
POLYGONS = []

LENGTHS = {}

for i, keys in enumerate(beams):
    plane = planes[i]

    points = []
    for key in keys:
        a = shell.vertex_coordinates(key)
        r = shell.get_vertex_attributes(key, ['rx', 'ry', 'rz'])
        b = add_vectors(a, r)

        line = a, b
        x = intersection_line_plane(line, plane)
        direction = normalize_vector(subtract_vectors(a, x))

        points.append(x)
        POINTS.append({
            'pos'   : x,
            'color' : (0, 0, 255),
            'name'  : "{}.{}.anchor".format(shell.name, key)
        })

        x0 = x
        x1 = add_vectors(x0, scale_vector(direction, 0.040))
        x2 = add_vectors(x1, scale_vector(direction, 0.300))
        x3 = add_vectors(x2, scale_vector(direction, 0.100))
        x4 = a

        LENGTHS[key] = [
            0.1,
            distance_point_point(x2, x3),
            distance_point_point(x3, x4),
        ]

        LINES.append({
            'start' : x0,
            'end'   : x1,
            'color' : (0, 0, 0),
            'name'  : "{}.anchor-ring".format(key)
        })
        LINES.append({
            'start' : x1,
            'end'   : x2,
            'color' : (255, 255, 255),
            'name'  : "{}.turn-buckle".format(key)
        })
        LINES.append({
            'start' : x2,
            'end'   : x3,
            'color' : (0, 0, 0),
            'name'  : "{}.sock".format(key)
        })
        LINES.append({
            'start' : x3,
            'end'   : x4,
            'color' : (255, 255, 255),
            'name'  : "{}.extra".format(key)
        })

    result = numerical.pca_numpy(points)

    origin = result[0][0]
    xaxis  = normalize_vector(result[1][0])
    yaxis  = normalize_vector(result[1][1])
    zaxis  = normalize_vector(result[1][2])
    frame = Frame(origin, xaxis, yaxis)

    if yaxis[2] > 0:
        offset = scale_vector(zaxis, -0.050)
    else:
        offset = scale_vector(zaxis, +0.050)

    points_xy = [frame.represent_point_in_local_coordinates(point) for point in points]
    box_xy = bounding_box_xy(points_xy)
    box = [frame.represent_point_in_global_coordinates(corner_xy)[:] for corner_xy in box_xy]
    box1 = offset_polygon(box, -0.025)
    box2 = [add_vectors(point, offset) for point in box1]

    POLYGONS.append({
        'points': box1 + box1[:1],
        'color' : (0, 255, 255),
    })
    POLYGONS.append({
        'points': box2 + box2[:1],
        'color' : (0, 0, 255),
    })

    LINES += framelines(origin, xaxis, yaxis, zaxis, 'box')

# ==============================================================================
# Export
# ==============================================================================

wb = Workbook()

boundary = set(shell.vertices_on_boundary())

beamS = [116, 261, 282, 60, 79, 52, 260, 5]
beamW = [57, 4, 259, 278, 54]
beamN = [268, 258, 115, 106, 114, 269, 281, 145]
beamSE = [146, 270, 117, 107]
beamNE = [8, 144, 272, 175]

ws_S = wb.active
ws_S.title = "SOUTH"
ws_W = wb.create_sheet()
ws_W.title = "WEST"
ws_N = wb.create_sheet()
ws_N.title = "NORTH"
ws_E = wb.create_sheet()
ws_E.title = "EAST"

append_to_sheet(ws_S, beamS, beamN)
append_to_sheet(ws_W, beamW, [])
append_to_sheet(ws_N, beamN, beamS)
append_to_sheet(ws_E, beamSE, beamNE)

wb.save(FILE_O)

# ==============================================================================
# Visualize
# ==============================================================================

# artist = ShellArtist(shell)

# artist.layer = "Scaffolding::Points"
# artist.clear_layer()
# artist.draw_points(POINTS)

# artist.layer = "Scaffolding::Cables"
# artist.clear_layer()
# artist.draw_lines(LINES)

# artist.layer = "Scaffolding::Beams"
# artist.clear_layer()
# artist.draw_polygons(POLYGONS)

